"""
Excel Export Service
--------------------
Generates a styled .xlsx file from the orders stored in PostgreSQL.
No Google Sheets dependency required.
"""

import io
import logging
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.db.models import Order, DHAKA_TZ

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_PAID_FILL     = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_PENDING_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_DATA_FONT     = Font(name="Calibri", size=10)
_CENTER        = Alignment(horizontal="center", vertical="center")
_LEFT          = Alignment(horizontal="left",   vertical="center")
_THIN          = Side(style="thin", color="CCCCCC")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

HEADERS = ["Order ID", "Product", "Qty", "Price", "Platform", "Timestamp", "Status", "Phone"]


async def generate_orders_excel(session: AsyncSession, days: int = 0) -> bytes:
    """
    Query orders from DB and return a styled .xlsx file as bytes.
    days=0 → all orders; days=N → last N days.
    """
    now = datetime.now(DHAKA_TZ).replace(tzinfo=None)
    stmt = select(Order).order_by(Order.timestamp.desc())
    if days > 0:
        stmt = stmt.where(Order.timestamp >= now - timedelta(days=days))

    result = await session.execute(stmt)
    orders = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales_Transaction_Log"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, order in enumerate(orders, 2):
        status_val  = order.payment_status.value if order.payment_status else ""
        platform_val = order.platform.value if order.platform else ""
        ts_val = order.timestamp.strftime("%Y-%m-%d %I:%M %p") if order.timestamp else ""
        phone_val = order.phone_number or ""

        row_data = [
            order.order_id,
            order.product_name,
            order.quantity,
            float(order.price),
            platform_val,
            ts_val,
            status_val,
            phone_val,
        ]

        # Pick row fill based on payment status
        row_fill = _PAID_FILL if status_val == "PAID" else _PENDING_FILL

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = _DATA_FONT
            cell.border    = _BORDER
            cell.alignment = _CENTER if col_idx in (3, 4, 7) else _LEFT
            # Colour the row subtly by status
            cell.fill = row_fill

        # Keep phone as text to preserve leading zeros
        if phone_val:
            ws.cell(row=row_idx, column=8).number_format = "@"

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [14, 28, 6, 10, 12, 22, 10, 16]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Summary footer ────────────────────────────────────────────────────────
    if orders:
        footer_row = len(orders) + 3
        ws.cell(row=footer_row, column=1, value="Total Orders").font = Font(bold=True)
        ws.cell(row=footer_row, column=2, value=len(orders))
        ws.cell(row=footer_row, column=3, value="Total Revenue").font = Font(bold=True)
        total_rev = sum(float(o.price) for o in orders)
        ws.cell(row=footer_row, column=4, value=round(total_rev, 2))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(f"Excel export generated: {len(orders)} orders")
    return buf.getvalue()
